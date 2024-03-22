import openpyxl
import csv
import excel_or_csv as ec

def csv_to_excel(input_csv_path, output_excel_path):
    """ csvファイルをExcelファイルに変換する関数 """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    # CSVファイルを開き、行ごとにExcelシートに書き込む
    with open(input_csv_path, 'r', newline='', encoding='utf-8') as csvfile:
        csv_reader = csv.reader(csvfile)
        for row_index, row in enumerate(csv_reader, start=1):
            for col_index, value in enumerate(row, start=1):
                sheet.cell(row=row_index, column=col_index, value=value)
    # Excelファイルに保存
    workbook.save(output_excel_path)

def list_to_csv(to_csv_list: list , csv_path: str = "output.csv"):
    """ 多次元リストのデータをcsvファイルに保存する関数 """
    with open(csv_path, 'w' , encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerows(to_csv_list)


def csv_to_list(csv_path: str = "output.csv"):
    """ 多次元データを含むcsvからリストに変換 """
    data_list = []
    with open(csv_path, 'r' , encoding="utf-8-sig") as file:
        csv_reader = csv.reader(file)
        for row in csv_reader:
            data_list.append(row)
    return data_list


def excel_to_list(input_excel_path: str = "input.xlsx"):
    workbook = openpyxl.load_workbook(input_excel_path)
    sheet = workbook.active
    row_num = sheet.max_row
    col_num = sheet.max_column
    data_list = []
    for row in range(1, row_num+1):
        row_data = []
        for col in range(1, col_num+1):
            cell_value = sheet.cell(row=row, column=col).value
            row_data.append(cell_value)
        data_list.append(row_data)
    return data_list

def list_to_excel(to_excel_list: list , output_excel_path: str = "output.xlsx"):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    # 多次元リストのサイズを取得(行ごとで列数に違いがあることを考慮)
    row_num , col_num = len(to_excel_list) , 0
    for row in range(row_num):
        predict_col = len(to_excel_list[row])
        if predict_col > col_num:
            col_num = predict_col
    for row in range(row_num):
        for col in range(col_num):
            try:
                sheet.cell(row=row+1, column=col+1).value = to_excel_list[row][col]
            except IndexError:
                pass
    workbook.save(output_excel_path)
    


def mail_list_from_excel(mail_excel_path):
    """ Excelファイルからメールのリストを取得する関数 """
    mail_list = []
    workbook = openpyxl.load_workbook(mail_excel_path)
    sheet = workbook.active
    receive_email_number = 100
    for index in range(receive_email_number):
        mail = sheet.cell(row = index + 2 , column = 3).value
        # emailかどうかを判定（「@」「.」の有無）
        if mail is not None:
            if '@' in mail and '.' in mail:
                mail_list.append(mail)
        else:
            break
    # ccのメールのリストを取得
    cc_mail_list = []
    for index in range(len(mail_list)):
        # D列以降を判定
        cc_mail_row_list = []
        for col in range(receive_email_number):
            cc_mail = sheet.cell(row = index + 2 , column = 4 + col).value
            # emailかどうかを判定（「@」「.」の有無）
            if cc_mail is not None:
                if '@' in cc_mail and '.' in cc_mail:
                    cc_mail_row_list.append(cc_mail)
            else:
                break
        cc_mail_list.append(cc_mail_row_list)
    # 送信元メールアドレスとアプリパスワードを取得
    from_email = sheet.cell(row = 2 , column = 1).value
    from_email_smtp_password = sheet.cell(row = 2 , column = 2).value
    return mail_list , cc_mail_list , from_email , from_email_smtp_password




def replace_excel_invalid_characters(input_string):
    """ 文字列の中から、Excelのシート名には使用できない文字を排除する関数 """
    invalid_characters = ['/', '\\', '?', '*', '[', ']', ':', '<', '>', '"', '|', '?', '*']
    for char in invalid_characters:
        input_string = input_string.replace(char, ' ')
    return input_string




def many_list_to_excel(
        many_list: list , input_excel_path , 
        output_excel_path: str = "output.xlsx" , sheet_name_list = []
    ):
    """ 多次元リストを複数含むリストをExcelファイルの複数のシートにそれぞれ転記する関数 """
    # workbook = openpyxl.Workbook()
    workbook = openpyxl.load_workbook(input_excel_path)
    # コピー元のシートを定義
    from_sheet_name = "input"
    from_sheet = workbook[from_sheet_name]

    for sheet_index, to_excel_list in enumerate(many_list):
        print(f"sheet_index : {sheet_index}")
        print(to_excel_list)
        print("\n")
        if sheet_name_list != []:
            to_sheet_name = f"{sheet_name_list[sheet_index]}"
            to_sheet_name = replace_excel_invalid_characters(to_sheet_name)
            # 名前を指定してコピー
            to_sheet = workbook.copy_worksheet(from_sheet)
            to_sheet.title = to_sheet_name
        else:
            to_sheet_name = f"Sheet{sheet_index + 1}"
            to_sheet_name = replace_excel_invalid_characters(to_sheet_name)
            # 名前を指定してコピー
            to_sheet = workbook.copy_worksheet(from_sheet)
            to_sheet.title = to_sheet_name

            
        """ レインズの取得結果を紐付け """
        # 多次元リストのサイズを取得(行ごとで列数に違いがあることを考慮)
        row_num, col_num = len(to_excel_list), 0
        for row in range(row_num):
            predict_col = len(to_excel_list[row])
            if predict_col > col_num:
                col_num = predict_col
        
        # 取得結果の貼り付け
        if len(to_excel_list) != 1:
            for row in range(5, row_num):  # 7行目以降のみループ
                for col in range(col_num):
                    try:
                        to_sheet.cell(row = row + 2, column = col + 1).value = to_excel_list[row][col]
                    except IndexError:
                        pass
        else:
            for row in range(row_num):
                for col in range(col_num):
                    try:
                        to_sheet.cell(row=row+7, column=col+1).value = to_excel_list[row][col]
                    except IndexError:
                        pass


    # 最初のデフォルトのシートを削除
    workbook.save(output_excel_path)






def get_search_option_from_excel(input_excel_path):
    """ Excelファイルから多次元リストを抽出する関数 """
    workbook = openpyxl.load_workbook(input_excel_path)
    sheet = workbook.active
    
    index_of_solding_requirement_list = []
    index_of_rental_requirement_list = []
    # 2行目のB列以降のデータを取得してリストに格納
    for cell in sheet[2][1:]:
        if cell == None:
            cell = 0
        index_of_solding_requirement_list.append(cell.value)
    # 3行目のB列以降のデータを取得してリストに格納
    for cell in sheet[3][1:]:
        if cell == None:
            cell = 0
        index_of_rental_requirement_list.append(cell.value)
    return index_of_solding_requirement_list, index_of_rental_requirement_list




def get_search_option(input_csv_path):
    """ 定期実行ツールがcsvファイルから検索方法と条件を取得する関数 """
    search_option_list = csv_to_list(input_csv_path)
    search_method_value = search_option_list[1][0]
    search_requirement = int( search_option_list[1][1] )
    return search_method_value , search_requirement