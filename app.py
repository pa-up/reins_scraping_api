from fastapi import FastAPI
from pydantic import BaseModel
import numpy as np
import re
import time
import os
import csv
import boto3
import openpyxl

from bs4 import BeautifulSoup
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait

import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication


static_path = "static"
# 入力パス
mail_excel_path = static_path + "/input_excel/email_pw.xlsx"
search_method_csv_path = static_path + "/csv/search_method.csv"
# 出力パス
output_reins_excel_path = static_path + "/output_excel/output_reins.xlsx"
log_txt_path = static_path + "/log/log.txt"

# 環境変数の取得
user_id , password = os.environ.get('SECRET_USER_ID') , os.environ.get('SECRET_PASSWORD')
s3_accesskey , s3_secretkey = os.environ.get('S3_ACCESSKEY') , os.environ.get('S3_SECRETKEY')
s3_bucket_name = os.environ.get('S3_BUCKET_NAME')


# s3の定義
s3_region = "ap-northeast-1"   # 東京(アジアパシフィック)：ap-northeast-1


def send_py_gmail(
    message_subject , message_body , from_email_smtp_password ,
    from_email , to_email , cc_mail_row_list = [] , file_path = "",
):
    """ メールを送信する関数 """
    msg = MIMEMultipart()
    msg['To'] = to_email
    msg['From'] = from_email
    if cc_mail_row_list !=[]:
        msg['Cc'] = ",".join(cc_mail_row_list)
    msg['Subject'] = message_subject
    msg.attach(MIMEText(message_body))
    # ファイルをメールに添付
    file_name = os.path.basename(file_path)
    with open(file_path , "rb") as f:
        attachment = MIMEApplication(f.read())
    attachment.add_header("Content-Disposition", "attachment", filename = file_name)
    msg.attach(attachment)
    # サーバーを指定しメールを送信
    smtp_server = "smtp.gmail.com"
    smtp_port = 587
    server = smtplib.SMTP(smtp_server, smtp_port)
    server.starttls()
    server.login(from_email, from_email_smtp_password)
    server.send_message(msg)
    server.quit()


def mail_list_from_excel(mail_excel_path):
    """ Excelファイルからメールのリストを取得する関数 """
    mail_list = []
    workbook = openpyxl.load_workbook(mail_excel_path)
    sheet = workbook.active
    receive_email_number = 100
    for index in range(receive_email_number):
        mail = sheet.cell(row = index + 2 , column = 3).value
        # emailかどうかを判定（「@」「.」の有無）
        if mail is not None:
            if '@' in mail and '.' in mail:
                mail_list.append(mail)
        else:
            break
    # ccのメールのリストを取得
    cc_mail_list = []
    for index in range(len(mail_list)):
        # D列以降を判定
        cc_mail_row_list = []
        for col in range(receive_email_number):
            cc_mail = sheet.cell(row = index + 2 , column = 4 + col).value
            # emailかどうかを判定（「@」「.」の有無）
            if cc_mail is not None:
                if '@' in cc_mail and '.' in cc_mail:
                    cc_mail_row_list.append(cc_mail)
            else:
                break
        cc_mail_list.append(cc_mail_row_list)
    # 送信元メールアドレスとアプリパスワードを取得
    from_email = sheet.cell(row = 2 , column = 1).value
    from_email_smtp_password = sheet.cell(row = 2 , column = 2).value
    return mail_list , cc_mail_list , from_email , from_email_smtp_password



def html_table_tag_to_csv_list(table_tag_str: str, header_exist: bool = True):
    table_soup = BeautifulSoup(table_tag_str, 'html.parser')
    rows = []
    if header_exist:
        for tr in table_soup.find_all('tr'):
            cols = [] 
            for td in tr.find_all(['td', 'th']):
                cols.append(td.text.strip())
            rows.append(cols)
    else:
        for tbody in table_soup.find_all('tbody'):
            for tr in tbody.find_all('tr'):
                cols = [td.text.strip() for td in tr.find_all(['td', 'th'])]
                rows.append(cols)
    return rows



def browser_setup(browse_visually = "no"):
    """ブラウザを起動する関数"""
    #ブラウザの設定
    options = webdriver.ChromeOptions()
    if browse_visually == "no":
        options.add_argument('--headless')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    #ブラウザの起動
    browser = webdriver.Chrome(options=options , service=ChromeService(ChromeDriverManager().install()))
    browser.implicitly_wait(3)
    return browser



class Reins_Scraper:
    def __init__(self, driver: WebDriverWait):
        self.driver = driver
        self.wait_driver = WebDriverWait(driver, 5)
    
    def login_reins(self, user_id: str , password: str ,):
        # ログインボタンをクリック
        login_button = self.wait_driver.until(EC.element_to_be_clickable((By.ID, "login-button")))
        login_button.click()

        # フォームにログイン認証情報を入力
        user_id_form = self.wait_driver.until(EC.presence_of_element_located((By.CSS_SELECTOR, "input[type='text']")))
        user_id_form.send_keys(user_id)
        time.sleep(0.5)
        password_form = self.wait_driver.until(EC.presence_of_element_located((By.CSS_SELECTOR, "input[type='password']")))
        password_form.send_keys(password)
        time.sleep(0.5)
        rule_element = self.wait_driver.until(EC.presence_of_element_located((By.XPATH, "//input[@type='checkbox' and contains(following-sibling::label, 'ガイドライン')]")))
        rule_checkbox_form = rule_element.find_element(By.XPATH, "./following-sibling::label")
        rule_checkbox_form.click()
        time.sleep(3)
        login_button = self.wait_driver.until(EC.presence_of_element_located((By.XPATH, "//button[contains(text(), 'ログイン')]")))
        login_button.click()
        time.sleep(3)

    def get_solding_or_rental_option(self):
        # ボタン「売買 物件検索」をクリック
        sold_building_search_button = self.wait_driver.until(EC.presence_of_element_located((By.XPATH, "//button[contains(text(), '売買') and contains(text(), '物件検索')]")))
        sold_building_search_button.click()
        time.sleep(1)
        # 検索条件を取得
        display_search_method_link = self.wait_driver.until(EC.presence_of_element_located((By.XPATH, "(//div[@class='card p-card'])[1]"))).find_element(By.XPATH, ".//a[contains(span, '検索条件を表示')]")
        display_search_method_link.click()
        time.sleep(1)
        # 検索条件のリストを取得
        select_element = self.wait_driver.until(EC.presence_of_element_located((By.XPATH, "//div[@class='p-selectbox']//select")))
        search_method_element_list = select_element.find_elements(By.TAG_NAME, "option")
        solding_search_method_list = []
        for search_method_element in search_method_element_list:
            solding_search_method_list.append( search_method_element.text )
        # 前のページに戻る
        self.driver.back()

        # ボタン「売買 物件検索」をクリック
        rental_building_search_button = self.wait_driver.until(EC.presence_of_element_located((By.XPATH, "//button[contains(text(), '賃貸') and contains(text(), '物件検索')]")))
        rental_building_search_button.click()
        time.sleep(1)
        # 検索条件を取得
        display_search_method_link = self.wait_driver.until(EC.presence_of_element_located((By.XPATH, "(//div[@class='card p-card'])[1]"))).find_element(By.XPATH, ".//a[contains(span, '検索条件を表示')]")
        display_search_method_link.click()
        time.sleep(1)
        # 検索条件のリストを取得
        select_element = self.wait_driver.until(EC.presence_of_element_located((By.XPATH, "//div[@class='p-selectbox']//select")))
        search_method_element_list = select_element.find_elements(By.TAG_NAME, "option")
        rental_search_method_list = []
        for search_method_element in search_method_element_list:
            rental_search_method_list.append( search_method_element.text )
        # 前のページに戻る
        self.driver.back()
        time.sleep(2)
        return solding_search_method_list , rental_search_method_list
        
    def scraping_solding_list(self , search_method_value: str , index_of_search_requirement: int):
        # 選択された検索方法をクリック
        if search_method_value == "search_solding":
            building_search_button = self.wait_driver.until(EC.presence_of_element_located((By.XPATH, "//button[contains(text(), '売買') and contains(text(), '物件検索')]")))
            building_search_button.click()
            time.sleep(1)
        else:
            building_search_button = self.wait_driver.until(EC.presence_of_element_located((By.XPATH, "//button[contains(text(), '賃貸') and contains(text(), '物件検索')]")))
            building_search_button.click()
            time.sleep(1)

        # 売買検索条件を選択
        display_search_method_link = self.wait_driver.until(EC.presence_of_element_located((By.XPATH, "(//div[@class='card p-card'])[1]"))).find_element(By.XPATH, ".//a[contains(span, '検索条件を表示')]")
        display_search_method_link.click()
        time.sleep(1)
        choice_search_method = Select(self.wait_driver.until(EC.presence_of_element_located((By.XPATH, "//div[@class='p-selectbox']//select"))))
        choice_search_method.select_by_index(index_of_search_requirement)
        get_button = self.wait_driver.until(EC.presence_of_element_located((By.XPATH, "//button[contains(text(), '読込')]")))
        get_button.click()
        time.sleep(1)
        time.sleep(0.5)

        # 検索条件が存在するか判定
        exist_search_requirement_sentence = self.wait_driver.until(EC.presence_of_element_located((By.CSS_SELECTOR, '[class*="modal"]'))).text
        if "エラー" in exist_search_requirement_sentence:
            to_csv_list = False
            self.driver.quit()
            return to_csv_list
        
        ok_button = self.wait_driver.until(EC.presence_of_element_located((By.XPATH, "//button[contains(text(), 'OK')]")))
        ok_button.click()
        time.sleep(1)

        # 検索条件に基づいて検索実行
        search_button = self.wait_driver.until(EC.presence_of_element_located((By.XPATH, "//div[@class='p-frame-bottom']//button[contains(text(), '検索')]")))
        search_button.click()
        time.sleep(2)

        # 物件リストが何ページあるかを判定
        time.sleep(2)
        page_count_info = self.wait_driver.until(EC.presence_of_element_located((By.CSS_SELECTOR, "div.card-header"))).text
        match = re.search(r'(\d+)件', page_count_info)
        total_number = int( match.group(1) )
        left_page_count = total_number / 50 

        # リストを取得
        loop_count = 0
        all_list = []
        while True:
            # 印刷表示ボタンをクリック
            print_button = self.wait_driver.until(EC.presence_of_element_located((By.XPATH, "//button[contains(text(), '印刷')]")))
            print_button.click()
            time.sleep(2)
            
            # 現在のページのHTML要素を取得
            table_tag_str = self.wait_driver.until(EC.presence_of_element_located((By.TAG_NAME, "table"))).get_attribute('outerHTML')
            # tableタグの要素を多次元リストに変換
            if loop_count == 0:
                header_exist = True
            else:
                header_exist = False
            loop_count += 1

            to_csv_list = html_table_tag_to_csv_list(
                table_tag_str = table_tag_str , header_exist = header_exist ,
            )
            all_list.append(to_csv_list)

            if left_page_count >= 1:
                left_page_count -= 1
                # リストの表示ページへ戻る
                back_button = self.wait_driver.until(EC.element_to_be_clickable((By.CLASS_NAME, 'p-frame-backer')))
                back_button.click()
                time.sleep(2)
                # 次のリストを表示させるボタンをクリック
                next_list_button = self.wait_driver.until(EC.element_to_be_clickable((By.CSS_SELECTOR, 'li.page-item > button > span.p-pagination-next-icon')))
                next_list_button.click()
                time.sleep(2)

            else:
                break

        self.driver.quit()
        
        # 全ての多次元リストを連結
        to_csv_list = []
        for loop in range( len(all_list) ):
            to_csv_list.extend( all_list[loop] )    
        
        return to_csv_list
    

def remove_non_number(text):
    divided_number = re.findall(r'\d+', text)  # 文字列から数字にマッチするものをリストとして取得
    integrate_only_number = ''.join(divided_number)
    integrate_only_number = re.sub(r'\D', '', text)  # 元の文字列から数字以外を削除＝数字を抽出
    return divided_number , integrate_only_number

def csv_to_excel(input_csv_path, output_excel_path):
    """ csvファイルをExcelファイルに変換する関数 """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    # CSVファイルを開き、行ごとにExcelシートに書き込む
    with open(input_csv_path, 'r', newline='', encoding='utf-8') as csvfile:
        csv_reader = csv.reader(csvfile)
        for row_index, row in enumerate(csv_reader, start=1):
            for col_index, value in enumerate(row, start=1):
                sheet.cell(row=row_index, column=col_index, value=value)
    # Excelファイルに保存
    workbook.save(output_excel_path)

def list_to_csv(to_csv_list: list , csv_path: str = "output.csv"):
    """ 多次元リストのデータをcsvファイルに保存する関数 """
    with open(csv_path, 'w' , encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerows(to_csv_list)


def csv_to_list(csv_path: str = "output.csv"):
    """ 多次元データを含むcsvからリストに変換 """
    data_list = []
    with open(csv_path, 'r' , encoding="utf-8-sig") as file:
        csv_reader = csv.reader(file)
        for row in csv_reader:
            data_list.append(row)
    return data_list


def excel_to_list(input_excel_path: str = "input.xlsx"):
    workbook = openpyxl.load_workbook(input_excel_path)
    log_txt.add_log_txt("Excelのワークブック起動完了 : workbook = openpyxl.load_workbook()")
    sheet = workbook.active
    log_txt.add_log_txt("ワークブックのアクティブ化完了 : sheet = workbook.active")
    row_num = sheet.max_row
    log_txt.add_log_txt(f"row_num : {row_num}")
    col_num = sheet.max_column
    log_txt.add_log_txt(f"col_num : {col_num}")
    data_list = []
    for row in range(1, row_num+1):
        row_data = []
        for col in range(1, col_num+1):
            cell_value = sheet.cell(row=row, column=col).value
            log_txt.add_log_txt(f"cell_value : {cell_value}")
            log_txt.add_log_txt(f"row , col : {row} , {col} \n")
            row_data.append(cell_value)
        data_list.append(row_data)
    log_txt.add_log_txt("セルの編集可能が証明 : cell_value = sheet.cell(row=row, column=col).value")
    return data_list

def list_to_excel(to_excel_list: list , output_excel_path: str = "output.xlsx"):
    workbook = openpyxl.Workbook()
    log_txt.add_log_txt("Excelのワークブック起動完了 : workbook = openpyxl.load_workbook()")
    sheet = workbook.active
    log_txt.add_log_txt("ワークブックのアクティブ化完了 : sheet = workbook.active")
    # 多次元リストのサイズを取得(行ごとで列数に違いがあることを考慮)
    row_num , col_num = len(to_excel_list) , 0
    for row in range(row_num):
        predict_col = len(to_excel_list[row])
        if predict_col > col_num:
            col_num = predict_col
    log_txt.add_log_txt(f"row_num , col_num : {row_num} , {col_num}")
    for row in range(row_num):
        for col in range(col_num):
            try:
                log_txt.add_log_txt(f"pressed_cell_value : {to_excel_list[row][col]}")
                log_txt.add_log_txt(f"row , col : {row} , {col} \n \n")
                sheet.cell(row=row+1, column=col+1).value = to_excel_list[row][col]
            except IndexError:
                pass
    log_txt.add_log_txt("セルの編集可能が証明 : sheet.cell(row=row+1, column=col+1).value = to_excel_list[row][col]")
    workbook.save(output_excel_path)
    

def get_search_option(input_csv_path):
    """ 定期実行ツールがcsvファイルから検索方法と条件を取得する関数 """
    search_option_list = csv_to_list(input_csv_path)
    search_method_value = search_option_list[1][0]
    search_requirement = int( search_option_list[1][1] )
    return search_method_value , search_requirement






class ManipulateS3:
    def __init__(self , accesskey , secretkey , bucket_name , region = "ap-northeast-1"):
        self.region = region  # 東京(アジアパシフィック)：ap-northeast-1
        self.accesskey = accesskey
        self.secretkey = secretkey
        self.bucket_name = bucket_name
        self.s3 = boto3.client('s3', aws_access_key_id=self.accesskey, aws_secret_access_key=self.secretkey, region_name=self.region)
    
    def get_file_name_from_file_path(self , file_path):
        """ パスからファイル名のみを抽出する関数
            s3にフォルダを作成し、ファイルをアップロードする場合は、この関数を使わずに、file_pathにフォルダ名を含める
        """
        file_name_from_path = file_path[file_path.rfind('/') + 1 : ]  # ファイルパスからファイル名のみを抽出
        return file_name_from_path
    
    def s3_file_upload(self , file_path):
        """ s3の特定のバケットにファイルをアップロードし、そのファイルのURLも取得する関数
            s3上のファイルが同一のファイル名であれば、s3内で上書き保存される
        """
        key_name = self.get_file_name_from_file_path(file_path)
        # s3へファイルをアップロード
        self.s3.upload_file(file_path, self.bucket_name, key_name)
        # S3へアップロードしたファイルへのURLを取得する
        s3_url = self.s3.generate_presigned_url(
            ClientMethod='get_object',
            Params={'Bucket': self.bucket_name, 'Key': key_name},
            ExpiresIn=3600,
            HttpMethod='GET'
        )
        return s3_url
    
    def s3_file_download(self , local_upload_path):
        """ s3の特定のバケットからファイル名で検索し、一致するファイルをダウンロードする関数
            local_file_pathのファイル名はs3で取得予定のファイル名を同一にする
        """
        key_name = self.get_file_name_from_file_path(local_upload_path)
        print(f"key_name : {key_name}")
        self.s3.download_file(self.bucket_name, key_name, local_upload_path)




class logText:
    def __init__(self , log_txt_path) -> None:
        self.log_txt_path = log_txt_path
        # logの保存ファイルを空にする
        with open(self.log_txt_path, 'w') as file:
            file.write('')

    def add_log_txt(self , add_log_text):
        """ logを付け加える関数 """
        with open(self.log_txt_path, 'a') as file:
            file.write("\n" + add_log_text)
log_txt = logText(log_txt_path)


class RequestDataScraping(BaseModel):
    search_method_value: list
    index_of_search_requirement: int
    mail_list: list
    cc_mail_list: list
    from_email: str
    from_email_smtp_password: str

class RequestDataExcel(BaseModel):
    to_excel_list: list
    search_method: str
    search_requirement: str
    mail_list: list
    cc_mail_list: list
    from_email: str
    from_email_smtp_password: str

# app = FastAPI()
app = FastAPI(default_response_limit=1024 * 1024 * 10)  # 10MBに増量


@app.post("/")
def fast_api_scraping():
    # S3からメール情報や検索条件を取得し、静的フォルダに格納
    manipulate_s3 = ManipulateS3(
        region = "ap-northeast-1" ,
        accesskey = s3_accesskey ,
        secretkey = s3_secretkey ,
        bucket_name = s3_bucket_name ,
    )
    manipulate_s3.s3_file_download(local_upload_path = mail_excel_path)
    manipulate_s3.s3_file_download(local_upload_path = search_method_csv_path)
    time.sleep(2)

    # csvファイルから検索方法と検索条件を選択（将来的に別のWEBアプリでも編集可能）
    search_method_value , index_of_search_requirement = get_search_option(search_method_csv_path)
    # メールアドレスのリストをExcelから取得
    mail_list , cc_mail_list , from_email , from_email_smtp_password = mail_list_from_excel(mail_excel_path)

    try:
        log_txt.add_log_txt("S3 取得完了")
        # ページにアクセス
        searched_url = "https://system.reins.jp/"
        driver = browser_setup()
        reins_sraper = Reins_Scraper(driver)
        driver.get(searched_url)
        log_txt.add_log_txt("reinsサイトにアクセス完了")

        # ログイン突破
        reins_sraper.login_reins(user_id , password)
        log_txt.add_log_txt("ログイン成功")
        # REINS上で存在する検索方法と検索条件を全て取得（01〜50番号まであることを前提）
        solding_search_method_list , rental_search_method_list = reins_sraper.get_solding_or_rental_option()
        log_txt.add_log_txt(f"solding_search_method_list , rental_search_method_list : \n{solding_search_method_list} \n{rental_search_method_list}")
        # スクレイピング結果のリストを取得
        to_excel_list = reins_sraper.scraping_solding_list(search_method_value , index_of_search_requirement)
        log_txt.add_log_txt("スクレイピング結果のリストを取得 : 完了")

        # 検索方法と検索条件を文字列で取得
        if search_method_value == "search_solding":
            search_method = "売買検索"
            search_requirement = solding_search_method_list[index_of_search_requirement]
        else:
            search_method = "賃貸検索"
            search_requirement = rental_search_method_list[index_of_search_requirement]
        log_txt.add_log_txt("検索方法と検索条件を文字列で取得 : 完了")

    except:
        # メールの送信文
        message_subject = "REINSスクレイピング定期実行"
        message_body = f"""
            スクレイピングができませんでした。エラーが発生しました。
        """
        file_path = log_txt_path

        # 全てのメールにスクレイピング結果のExcelを送信
        for loop , to_email in enumerate(mail_list):
            cc_mail_row_list = cc_mail_list[loop]
            send_py_gmail(
                message_subject , message_body , from_email_smtp_password ,
                from_email , to_email , cc_mail_row_list = cc_mail_row_list ,
                file_path = file_path ,
            )

    return {
        "to_excel_list": to_excel_list ,
        "search_method" : search_method ,
        "search_requirement" : search_requirement ,
        "mail_list" : mail_list ,
        "cc_mail_list" : cc_mail_list ,
        "from_email" : from_email ,
        "from_email_smtp_password" : from_email_smtp_password ,
    }


@app.post("/excel")
def fast_api_excel(api_data_excel: RequestDataExcel):
    log_txt.add_log_txt("2つ目のAPI起動完了")
    to_excel_list = api_data_excel.to_excel_list
    search_method = api_data_excel.search_method
    search_requirement = api_data_excel.search_requirement

    mail_list = api_data_excel.mail_list
    cc_mail_list = api_data_excel.cc_mail_list
    from_email = api_data_excel.from_email
    from_email_smtp_password = api_data_excel.from_email_smtp_password

    try:
        # スクレイピング結果のリストをExcelファイルに保存
        list_to_excel(to_excel_list , output_reins_excel_path)
        ##### 最終的にはExcelの定型フォームに貼り付け
        log_txt.add_log_txt("スクレイピング結果をExcelファイルに変更 : 完了")
        
        # メールの送信文
        message_subject = "REINSスクレイピング定期実行"
        message_body = f"""
            REINSの定期日時スクレイピング結果のメールです。
            検索方法 : 「{search_method}」
            検索条件：「{search_requirement}」
            ※ 検索条件は「01」〜「50」の番号で指定されます

            スクレイピング結果は添付のExcelファイルをご覧ください。

            指定日時実行の検索条件を変更する際は、ツール「web_reins」で設定変更が可能です。
            変更後再度、cronでMac OS上の処理スケジュールを変更する必要があります。
            （※ cronの設定方法もツール「web_reins」でご確認いただけます。）
        """
        file_path = output_reins_excel_path
    except Exception as error_data:
        error_text = str(error_data)
        # メールの送信文
        message_subject = "REINSスクレイピング定期実行"
        message_body = f"""
            Excelファイル化ができませんでした。エラーが発生しました。
            ========================================
            検索方法 : 「{search_method}」
            検索条件：「{search_requirement}」
            ========================================
            エラーメッセージ :
            ----------------------------------------
            {error_text}
            ========================================

            ========================================
            REINSのExcelリスト :
            ----------------------------------------
            {to_excel_list}
            ========================================
        """
        file_path = log_txt_path

    # 全てのメールにスクレイピング結果のExcelを送信
    for loop , to_email in enumerate(mail_list):
        cc_mail_row_list = cc_mail_list[loop]
        send_py_gmail(
            message_subject , message_body , from_email_smtp_password ,
            from_email , to_email , cc_mail_row_list = cc_mail_row_list ,
            file_path = file_path ,
        )
    return {"message_body": message_body}


